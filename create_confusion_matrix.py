import argparse
import os
import json
import re
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
from collections import defaultdict
from openai import OpenAI

load_dotenv()

GEMINI_MODEL_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-3-flash-preview:generateContent"
)

DEFAULT_BATCH_SIZE = 5

# file to extract a dataframe from the excel file indicated as the input.
def read_excel(path):
    try:
        df = pd.read_excel(path).head(100)
        print(f"Read input file: {path}")
        return df
    except Exception as e:
        raise RuntimeError(f"Failed to read Excel file: {e}")

# write the confusion matrix to a new excel file with some basic formatting.
def write_confusion_excel(matrix_df, output_path):
    if os.path.exists(output_path):
        os.remove(output_path)
        print(f"Deleted existing file: {output_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Confusion Matrix"

    ws["A1"] = "Confusion Matrix"
    ws["A1"].font = Font(bold=True)

    ws.append([])

    for row in dataframe_to_rows(matrix_df, index=True, header=True):
        ws.append(row)

    wb.save(output_path)
    print(f"Wrote output file: {output_path}")

def extract_json(text):
    text = re.sub(r"```json|```", "", text).strip()
    return text

# batch the dataframe into smaller chunks to speed up processing and avoid hitting API limits.
def batch_dataframe(df, batch_size):
    for i in range(0, len(df), batch_size):
        yield df.iloc[i:i + batch_size]

# send a batch of data to GPT-4 and get true and predicted labels. Parse the response as a json.
def call_gpt(batch_df, prompt, properties, labels):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY environment variable not set")

    client = OpenAI(api_key=api_key)

    prompt = f"""
{prompt}

Data:
{batch_df.to_json(orient="records")}
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ],
            temperature=0,
            response_format = {
                "type": "json_schema",
                "json_schema": {
                    "name": "radiology_classification",
                    "schema": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "results": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "additionalProperties": False,
                                    "properties": properties,
                                    "required": labels
                                }
                            }
                        },
                        "required": ["results"]
                    }
                }
            },
        )

        raw = response.choices[0].message.content
        # print(f"Raw API response:\n{raw}")
        parsed = json.loads(raw)
        results = parsed["results"]

        print(f"Batch returned {len(results)} lines.")

        return results

    except Exception as e:
        raise RuntimeError(f"OpenAI API error: {e}")

# Function to extract ground truth labels from specific cells in the Excel file
def extract_ground_truth_labels(path, labels, cols):
    try:
        ground_truth_df = pd.read_excel(
            path, usecols=cols, skiprows=0, nrows=100
        )
        label_names = labels

        ground_truth_df.columns = label_names
        print("Extracted ground truth labels from input file.")
        return ground_truth_df
    except Exception as e:
        raise RuntimeError(f"Failed to extract ground truth labels: {e}")

# Function to load results from an Excel file instead of calling the LLM
def load_results_from_excel(file_path):
    try:
        results_df = pd.read_excel(file_path)
        print(f"Loaded results from: {file_path}")
        return results_df.to_dict(orient="records")
    except Exception as e:
        raise RuntimeError(f"Failed to load results from Excel file: {e}")

def to_binary(x):
    return 1 if x == "Abnormal" else 0

def extract_matching_prompt(input_file_name):
    json_path = os.path.join(os.path.dirname(__file__), "prompts.json")
    try:
        with open(json_path, "r") as file:
            data = json.load(file)
    except Exception as e:
        raise RuntimeError(f"Failed to load JSON file: {e}")

    # Navigate to the 'prompts' key in the JSON structure
    prompts = data.get("prompts", [])

    for entry in prompts:
        if entry.get("fileName") == input_file_name:
            return entry

    raise ValueError(f"No matching entry found for fileName: {input_file_name}")

# main function to read command line input, batch the data, and write the confusion matrix to a new excel file.
def main():
    parser = argparse.ArgumentParser(
        description="Generate a confusion matrix from Excel using Google Gemini"
    )
    parser.add_argument("input_file", help="Path to input Excel file")
    parser.add_argument(
        "--batch-size",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help="Number of rows per GPT request"
    )

    args = parser.parse_args()

    df = read_excel(args.input_file)

    matching_prompt = extract_matching_prompt(os.path.basename(args.input_file))
    prompt = matching_prompt.get("prompt", "")
    properties = matching_prompt.get("properties", {})
    labels = matching_prompt.get("labels", [])
    cols = matching_prompt.get("cols", "")

    all_results = []
    batch_index = 1
    for batch in batch_dataframe(df, args.batch_size):
        print(f"Processing batch {batch_index} of {len(batch)} rows...")
        # print(f"Batch input data:\n{batch}")


        for attempt in range(3):  # Retry up to 3 times, had trouble with gpt returning number of lines not matching input rows
            try:
                batch_results = call_gpt(batch, prompt, properties, labels)
                if len(batch_results) == len(batch):
                    break
                else:
                    print(f"Retrying batch {batch_index}, attempt {attempt + 1}")
            except Exception as e:
                print(f"Error on attempt {attempt + 1}: {e}")

        all_results.extend(batch_results)
        batch_index += 1

    # print(all_results)
    # all_results = load_results_from_excel("all_results.xlsx")

    # Save all_results to an Excel file
    all_results_df = pd.DataFrame(all_results)
    all_results_output_file = "all_results.xlsx"
    all_results_df.to_excel(all_results_output_file, index=False)
    print(f"All results saved to: {all_results_output_file}")

    # print(all_results)

    print(labels)

    ground_truth_labels = extract_ground_truth_labels(args.input_file, labels, cols)

    conf_matrix = {
        label: {"TP": 0, "TN": 0, "FP": 0, "FN": 0}
        for label in labels
    }
    global_counts = {"TP": 0, "TN": 0, "FP": 0, "FN": 0}

    for i in range(len(all_results)):

        ai_case = all_results[i]
        gt_row = ground_truth_labels.iloc[i]

        for label in labels:

            y_pred = to_binary(ai_case[label])
            y_true = to_binary(gt_row[label])

            if y_true == 1 and y_pred == 1:
                conf_matrix[label]["TP"] += 1
                global_counts["TP"] += 1

            elif y_true == 0 and y_pred == 0:
                conf_matrix[label]["TN"] += 1
                global_counts["TN"] += 1

            elif y_true == 0 and y_pred == 1:
                conf_matrix[label]["FP"] += 1
                global_counts["FP"] += 1

            elif y_true == 1 and y_pred == 0:
                conf_matrix[label]["FN"] += 1
                global_counts["FN"] += 1

    # Calculate sensitivity and specificity for each label
    metrics = {}
    for label, counts in conf_matrix.items():
        tp = counts["TP"]
        tn = counts["TN"]
        fp = counts["FP"]
        fn = counts["FN"]

        sensitivity = tp / (tp + fn) if (tp + fn) > 0 else 0
        specificity = tn / (tn + fp) if (tn + fp) > 0 else 0
        accuracy = (tp + tn) / (tp + tn + fp + fn) if (tp + tn + fp + fn) > 0 else 0

        metrics[label] = {
            "Sensitivity": sensitivity,
            "Specificity": specificity,
            "Accuracy": accuracy
        }

    # print("Label-wise Sensitivity and Specificity:")
    # for label, metric in metrics.items():
    #     print(f"{label}: Sensitivity = {metric['Sensitivity']:.2f}, Specificity = {metric['Specificity']:.2f}")

    metrics_df = pd.DataFrame.from_dict(metrics, orient="index")
    metrics_df.reset_index(inplace=True)
    metrics_df.rename(columns={"index": "Label"}, inplace=True)

    conf_matrix_df = pd.DataFrame.from_dict(conf_matrix, orient="index")
    combined_df = conf_matrix_df.merge(metrics_df, left_index=True, right_on="Label")

    combined_output_file = "combined_results.xlsx"
    write_confusion_excel(combined_df, combined_output_file)

    print(f"Combined confusion matrix and metrics saved to: {combined_output_file}")

if __name__ == "__main__":
    main()